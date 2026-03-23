# src/fibot/analysis/show_results.py
# FiBot — Ergebnisanzeige und Live Signal-Check

import os
import sys
import json
import logging
import argparse
from datetime import date
from typing import Optional

import pandas as pd

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.append(os.path.join(PROJECT_ROOT, 'src'))

logging.basicConfig(level=logging.WARNING, format='%(levelname)s %(message)s')
logger = logging.getLogger(__name__)

RESULTS_DIR = os.path.join(PROJECT_ROOT, 'artifacts', 'results')
SETTINGS_FILE = os.path.join(PROJECT_ROOT, 'settings.json')

# ANSI Farben
GREEN  = '\033[0;32m'
YELLOW = '\033[1;33m'
RED    = '\033[0;31m'
CYAN   = '\033[0;36m'
BOLD   = '\033[1m'
NC     = '\033[0m'

# ---------------------------------------------------------------------------
# Modus 1: Einzel-Analyse — alle Configs isoliert testen (stbot-Style)
# ---------------------------------------------------------------------------

def run_all_configs_isolated(date_from: str, date_to: str, capital: float):
    from fibot.analysis.backtester import run_backtest, load_ohlcv

    if not os.path.isdir(CONFIGS_DIR):
        print(f"{RED}Kein Configs-Verzeichnis: {CONFIGS_DIR}{NC}")
        return

    cfg_files = sorted(f for f in os.listdir(CONFIGS_DIR)
                       if f.startswith('config_') and f.endswith('.json'))
    if not cfg_files:
        print(f"{YELLOW}Keine Configs gefunden. Erst run_pipeline.sh ausführen.{NC}")
        return

    print(f"\n--- FiBot Ergebnis-Analyse (Einzel-Modus) ---")
    print(f"Zeitraum: {date_from} bis {date_to} | Startkapital: {capital:.0f} USDT\n")

    results = []
    for fname in cfg_files:
        cfg_path = os.path.join(CONFIGS_DIR, fname)
        try:
            with open(cfg_path) as f:
                config = json.load(f)
        except Exception:
            continue

        symbol    = config.get('market', {}).get('symbol', '')
        timeframe = config.get('market', {}).get('timeframe', '')
        if not symbol or not timeframe:
            continue

        print(f"Analysiere Ergebnisse für: {fname}...")
        df = load_ohlcv(symbol, timeframe, date_from, date_to)
        if df.empty or len(df) < 50:
            print(f"  {YELLOW}Keine Daten — übersprungen.{NC}")
            continue

        result = run_backtest(df, config, capital, symbol, timeframe)
        results.append({
            'symbol':    symbol,
            'timeframe': timeframe,
            'trades':    result.total_trades,
            'win_rate':  result.win_rate,
            'pnl_pct':   result.pnl_pct,
            'max_dd':    result.max_drawdown_pct,
            'end_cap':   result.end_capital,
        })

    if not results:
        print(f"{RED}Kein Backtest erfolgreich.{NC}")
        return

    W = 89
    print(f"\n{'='*W}")
    print(f"{'Zusammenfassung aller Einzelstrategien':^{W}}")
    print(f"{'='*W}")
    print(f"  {'Strategie':<22}  {'Trades':>6}  {'Win Rate %':>10}  {'PnL %':>7}  {'Max DD %':>8}  {'Endkapital':>10}")
    for r in sorted(results, key=lambda x: x['pnl_pct'], reverse=True):
        strat  = f"{r['symbol']} ({r['timeframe']})"
        color  = GREEN if r['pnl_pct'] >= 0 else RED
        dd_col = GREEN if r['max_dd'] <= 30 else RED
        print(f"  {strat:<22}  {r['trades']:>6}  {r['win_rate']:>10.2f}  "
              f"{color}{r['pnl_pct']:>7.2f}{NC}  "
              f"{dd_col}{r['max_dd']:>8.2f}{NC}  "
              f"{r['end_cap']:>10.2f}")
    print(f"{'='*W}")


# ---------------------------------------------------------------------------
# Modus 2: Manuelle Portfolio-Simulation — User wählt Configs
# ---------------------------------------------------------------------------

def run_manual_portfolio(filenames: list, date_from: str, date_to: str, capital: float):
    from fibot.analysis.backtester import run_backtest, load_ohlcv

    print(f"\n--- FiBot Manuelle Portfolio-Simulation ---")
    print(f"Zeitraum: {date_from} bis {date_to} | Startkapital: {capital:.0f} USDT\n")

    results = []
    for fname in filenames:
        cfg_path = os.path.join(CONFIGS_DIR, fname.strip())
        if not os.path.exists(cfg_path):
            print(f"  {RED}Config nicht gefunden: {fname}{NC}")
            continue
        try:
            with open(cfg_path) as f:
                config = json.load(f)
        except Exception:
            continue

        symbol    = config.get('market', {}).get('symbol', '')
        timeframe = config.get('market', {}).get('timeframe', '')
        if not symbol or not timeframe:
            continue

        print(f"Analysiere Ergebnisse für: {fname.strip()}...")
        df = load_ohlcv(symbol, timeframe, date_from, date_to)
        if df.empty or len(df) < 50:
            print(f"  {YELLOW}Keine Daten — übersprungen.{NC}")
            continue

        result = run_backtest(df, config, capital, symbol, timeframe)
        results.append({
            'symbol':    symbol,
            'timeframe': timeframe,
            'trades':    result.total_trades,
            'win_rate':  result.win_rate,
            'pnl_pct':   result.pnl_pct,
            'max_dd':    result.max_drawdown_pct,
            'end_cap':   result.end_capital,
        })

    if not results:
        print(f"{RED}Kein Backtest erfolgreich.{NC}")
        return

    W = 89
    print(f"\n{'='*W}")
    print(f"{'Manuelle Portfolio-Simulation':^{W}}")
    print(f"{'='*W}")
    print(f"  {'Strategie':<22}  {'Trades':>6}  {'Win Rate %':>10}  {'PnL %':>7}  {'Max DD %':>8}  {'Endkapital':>10}")
    for r in results:
        strat  = f"{r['symbol']} ({r['timeframe']})"
        color  = GREEN if r['pnl_pct'] >= 0 else RED
        dd_col = GREEN if r['max_dd'] <= 30 else RED
        print(f"  {strat:<22}  {r['trades']:>6}  {r['win_rate']:>10.2f}  "
              f"{color}{r['pnl_pct']:>7.2f}{NC}  "
              f"{dd_col}{r['max_dd']:>8.2f}{NC}  "
              f"{r['end_cap']:>10.2f}")

    # Kombiniertes Portfolio (Kapital / N)
    n        = len(results)
    per_cap  = capital / n
    port_end = sum(per_cap * (1 + r['pnl_pct'] / 100) for r in results)
    port_pnl = (port_end - capital) / capital * 100
    port_dd  = max(r['max_dd'] for r in results)
    col      = GREEN if port_pnl >= 0 else RED
    print(f"{'─'*W}")
    print(f"  Portfolio ({n} Strategie(n), je {per_cap:.2f} USDT):  "
          f"{col}{port_pnl:+.2f}%{NC}  |  End: {col}{port_end:.2f} USDT{NC}  |  MaxDD: {port_dd:.2f}%")
    print(f"{'='*W}")


# ---------------------------------------------------------------------------
# Modus 3: Portfolio-Optimierer — beste Coins/TFs für gegebene Randbedingungen
# ---------------------------------------------------------------------------

CONFIGS_DIR = os.path.join(PROJECT_ROOT, 'src', 'fibot', 'strategy', 'configs')
OPT_RESULTS  = os.path.join(PROJECT_ROOT, 'artifacts', 'results', 'optimization_results.json')


def run_portfolio_finder(capital: float, target_max_dd: float, min_wr: float,
                          start_date: str, end_date: str, auto: bool = False,
                          symbols: list | None = None,
                          configs: list | None = None):
    """
    Findet das optimale Fibonacci-Portfolio per Greedy-Algorithmus — stbot-Style.

    Schritt 1: Alle Configs backtesten (isoliert), nach MaxDD + MinWR filtern.
    Schritt 2: Beste Einzelstrategie als Startpunkt.
    Schritt 3: Greedy — iterativ echte Portfolio-Simulation (run_portfolio_simulation)
               pro Kandidat: gemeinsames Kapital, chronologisch, Margin-Check.
    Coin-Kollision: kein Coin doppelt (BTC 4h + BTC 1h = blockiert).
    """
    from fibot.analysis.backtester import run_backtest, load_ohlcv
    from fibot.strategy.fibonacci_logic import precompute_indicators, precompute_all_signals
    from fibot.analysis.portfolio_simulator import run_portfolio_simulation

    if not os.path.isdir(CONFIGS_DIR):
        print(f"{RED}Kein Configs-Verzeichnis: {CONFIGS_DIR}{NC}")
        return

    cfg_files = sorted(f for f in os.listdir(CONFIGS_DIR)
                       if f.startswith('config_') and f.endswith('.json'))
    if not cfg_files:
        print(f"{YELLOW}Keine Configs gefunden. Erst run_pipeline.sh ausführen.{NC}")
        return

    # Configs-Filter: exakte Dateiliste hat Vorrang, sonst settings.json aktive Strategien
    if configs:
        cfg_files = [f for f in configs if f in cfg_files]
        if not cfg_files:
            print(f"{RED}Keine der angegebenen Configs gefunden.{NC}")
            return
        print(f"  Configs-Filter aktiv: {', '.join(cfg_files)}")
    elif not symbols:
        # Automatisch aus settings.json filtern
        try:
            with open(SETTINGS_FILE) as _sf:
                _settings = json.load(_sf)
            _active = _settings.get('live_trading_settings', {}).get('active_strategies', [])
            if _active:
                _cfg_names = set()
                for _s in _active:
                    _sym = _s.get('symbol', '')
                    _tf  = _s.get('timeframe', '')
                    if _sym and _tf:
                        _safe = f"{_sym.replace('/', '').replace(':', '')}_{_tf}"
                        _cfg_names.add(f"config_{_safe}_fib.json")
                cfg_files = [f for f in cfg_files if f in _cfg_names]
                print(f"  Auto-Filter (settings.json): {', '.join(cfg_files)}")
        except Exception:
            pass
    if symbols:
        allowed = {s.upper().split('/')[0] for s in symbols}
        cfg_files = [f for f in cfg_files
                     if any(f.upper().startswith(f'CONFIG_{coin}') for coin in allowed)]
        if not cfg_files:
            print(f"{RED}Keine Configs für Symbole {allowed} gefunden.{NC}")
            return
        print(f"  Symbol-Filter aktiv: {', '.join(sorted(allowed))}")

    cond = f"Max DD <= {target_max_dd:.2f}%"
    if min_wr > 0:
        cond += f" & WR >= {min_wr:.1f}%"
    print(f"\n--- Starte automatische Portfolio-Optimierung (FiBot) mit {cond} & ohne Coin-Kollisionen ---")

    # ── Schritt 1: Daten laden, Signale vorberechnen, Einzel-Backtest ────────
    print(f"\n1/3: Analysiere Einzel-Performance & filtere nach {cond}...")
    print(f"     Zeitraum: {start_date} bis {end_date} | Startkapital: {capital:.0f} USDT\n")

    single_results  = []   # für Filterung
    strategies_data = {}   # für Portfolio-Simulator (df bereits precomputed)

    for fname in cfg_files:
        cfg_path = os.path.join(CONFIGS_DIR, fname)
        try:
            with open(cfg_path) as f:
                config = json.load(f)
        except Exception as e:
            print(f"  {RED}Lesefehler {fname}: {e}{NC}")
            continue

        symbol    = config.get('market', {}).get('symbol', '')
        timeframe = config.get('market', {}).get('timeframe', '')
        if not symbol or not timeframe:
            continue

        df = load_ohlcv(symbol, timeframe, start_date, end_date)
        if df.empty or len(df) < 50:
            print(f"  {YELLOW}Uebersprungen (keine Daten): {fname}{NC}")
            continue

        # Indikatoren + Signale vorberechnen (einmal, wird für Simulator wiederverwendet)
        df = precompute_indicators(df, config)
        df = precompute_all_signals(df, config)

        result = run_backtest(df, config, capital, symbol, timeframe)

        strategies_data[fname] = {
            'symbol':    symbol,
            'timeframe': timeframe,
            'df':        df,
            'config':    config,
        }
        single_results.append({
            'filename':   fname,
            'symbol':     symbol,
            'timeframe':  timeframe,
            'coin':       symbol.split('/')[0],
            'pnl_pct':    result.pnl_pct,
            'end_cap':    result.end_capital,
            'win_rate':   result.win_rate,
            'max_dd':     result.max_drawdown_pct,
            'trades':     result.total_trades,
            'trade_objs': result.trades,
        })

        ok      = result.max_drawdown_pct <= target_max_dd and result.win_rate >= min_wr
        dd_col  = GREEN if result.max_drawdown_pct <= target_max_dd else RED
        pnl_col = GREEN if result.pnl_pct >= 0 else RED
        print(f"  [{'OK' if ok else '--'}] {fname:<44}  "
              f"PnL {pnl_col}{result.pnl_pct:>+7.2f}%{NC}  "
              f"WR {result.win_rate:>5.1f}%  "
              f"DD {dd_col}{result.max_drawdown_pct:>6.2f}%{NC}  "
              f"Trades {result.total_trades:>4}")

    if not single_results:
        print(f"{RED}Kein Backtest erfolgreich. Abbruch.{NC}")
        return

    valid = [r for r in single_results
             if r['max_dd'] <= target_max_dd and r['win_rate'] >= min_wr]

    if not valid:
        print(f"\n{RED}Keine Einzelstrategie erfuellt {cond}.{NC}")
        min_dd = min(r['max_dd'] for r in single_results)
        print(f"  Niedrigster erreichbarer DD: {min_dd:.1f}%")
        print(f"  TIPP: Max Drawdown auf mindestens {int(min_dd) + 5}% erhoehen.")
        return

    # ── Schritt 2: Besten Einzelspieler wählen (validiert gegen Portfolio-Sim-DD) ──
    valid.sort(key=lambda x: x['end_cap'], reverse=True)

    # Aktuelle Portfolio-Performance per Simulation
    def _simulate(filenames: list) -> dict | None:
        data = {fn: strategies_data[fn] for fn in filenames if fn in strategies_data}
        return run_portfolio_simulation(capital, data, start_date, end_date)

    best_single  = None
    best_sim     = None
    best_end_cap = 0.0
    best_dd      = 0.0

    print(f"\n2/3: Suche Basis-Strategie (Portfolio-Sim-DD muss <= {target_max_dd:.2f}%)...")
    for candidate in valid:
        sim = _simulate([candidate['filename']])
        sim_dd = sim['max_drawdown_pct'] if sim else candidate['max_dd']
        dd_ok  = sim_dd <= target_max_dd
        cap    = sim['end_capital'] if sim else candidate['end_cap']
        col    = GREEN if dd_ok else RED
        print(f"  {'OK' if dd_ok else '--'} {candidate['filename']:<44}  "
              f"Sim: {cap:.2f} USDT  Sim-DD: {col}{sim_dd:.2f}%{NC}")
        if dd_ok and (best_single is None or cap > best_end_cap):
            best_single  = candidate
            best_sim     = sim
            best_end_cap = cap
            best_dd      = sim_dd

    if best_single is None:
        print(f"\n{RED}Keine Einzelstrategie erfuellt {cond} auch in der Portfolio-Simulation.{NC}")
        print(f"  TIPP: Max Drawdown erhoehen (Portfolio-Sim berechnet Mark-to-Market DD).")
        return

    print(f"\n     Beste Basis: {best_single['filename']}")
    print(f"     Einzel-Backtest: {best_single['end_cap']:.2f} USDT, Max DD: {best_single['max_dd']:.2f}%")
    print(f"     Portfolio-Simulation: {best_end_cap:.2f} USDT, Sim-DD: {best_dd:.2f}%")

    # ── Schritt 3: Greedy mit echter Portfolio-Simulation ────────────────────
    print(f"\n3/3: Suche beste Team-Kollegen (echte Portfolio-Simulation, gemeinsames Kapital)...")

    portfolio      = [best_single]
    used_coins     = {best_single['coin']}
    candidate_pool = [r for r in valid if r['filename'] != best_single['filename']]

    print(f"     Verbesserungen werden relativ zur Basis-Simulation bewertet.")

    while True:
        best_next    = None
        best_cap_w   = best_end_cap
        best_dd_w    = best_dd

        for candidate in candidate_pool:
            if candidate['coin'] in used_coins:
                continue   # Coin-Kollision

            # Schnell-Check: Einzel-DD des Kandidaten muss <= target
            if candidate['max_dd'] > target_max_dd:
                continue

            filenames = [r['filename'] for r in portfolio] + [candidate['filename']]
            sim = _simulate(filenames)
            if sim is None:
                continue
            if sim['liquidation_date'] is not None:
                continue
            if sim['max_drawdown_pct'] > target_max_dd:
                continue
            if sim['end_capital'] > best_cap_w:
                best_cap_w = sim['end_capital']
                best_dd_w  = sim['max_drawdown_pct']
                best_next  = candidate

        if best_next:
            portfolio.append(best_next)
            used_coins.add(best_next['coin'])
            candidate_pool.remove(best_next)
            best_end_cap = best_cap_w
            best_dd      = best_dd_w
            print(f"-> Fuege hinzu: {best_next['filename']}"
                  f"  (Neues Kapital: {best_cap_w:.2f} USDT, Max DD: {best_dd_w:.2f}%)")
        else:
            print("Keine weitere Verbesserung moeglich (echte Simulation, DD & Coin-Kollision). "
                  "Optimierung beendet.")
            break

    # ── Finale Simulation des optimalen Portfolios ────────────────────────────
    final_filenames = [r['filename'] for r in portfolio]
    final_sim = _simulate(final_filenames)

    if final_sim:
        port_end_cap = final_sim['end_capital']
        port_pnl_pct = final_sim['total_pnl_pct']
        port_max_dd  = final_sim['max_drawdown_pct']
        port_trades  = final_sim['trade_count']
        port_wr      = final_sim['win_rate']
        liquidated   = final_sim['liquidation_date'] is not None
    else:
        port_end_cap = best_end_cap
        port_pnl_pct = (best_end_cap - capital) / capital * 100
        port_max_dd  = best_dd
        port_trades  = sum(r['trades'] for r in portfolio)
        port_wr      = 0.0
        liquidated   = False

    pnl_col  = GREEN if port_pnl_pct >= 0 else RED
    n_strats = len(portfolio)

    print(f"\n{'='*55}")
    print(f"     Ergebnis der automatischen Portfolio-Optimierung")
    print(f"{'='*55}")
    print(f"Zeitraum: {start_date} bis {end_date}")
    print(f"Startkapital: {capital:.2f} USDT")
    print(f"Bedingung: Max Drawdown <= {target_max_dd:.2f}%"
          + (f"  |  WR >= {min_wr:.1f}%" if min_wr > 0 else ""))

    print(f"\nOptimales Portfolio gefunden ({n_strats} Strategie(n)):")
    for r in portfolio:
        print(f"  - {r['filename']}")

    print(f"\n--- Simulierte Performance dieses Portfolios ---")
    print(f"Endkapital:         {pnl_col}{port_end_cap:.2f} USDT{NC}")
    print(f"Gesamt PnL:         {pnl_col}{port_end_cap - capital:+.2f} USDT "
          f"({port_pnl_pct:.2f}%){NC}")
    print(f"Trades gesamt:      {port_trades}  |  Win-Rate: {port_wr:.1f}%")
    print(f"Portfolio Max DD:   {port_max_dd:.2f}%")
    print(f"Liquidiert:         {'JA' if liquidated else 'NEIN'}")
    print(f"{'='*55}\n")

    # --- Ergebnis speichern ---
    portfolio_set = set(final_filenames)
    os.makedirs(os.path.dirname(OPT_RESULTS), exist_ok=True)
    with open(OPT_RESULTS, 'w') as f:
        json.dump({
            'optimal_portfolio':  final_filenames,
            'end_capital':        port_end_cap,
            'total_pnl_pct':      port_pnl_pct,
            'max_drawdown_pct':   port_max_dd,
            'trade_count':        port_trades,
            'win_rate':           port_wr,
            'all_results': [
                {
                    'filename':   r['filename'],
                    'symbol':     r['symbol'],
                    'timeframe':  r['timeframe'],
                    'pnl_pct':    r['pnl_pct'],
                    'end_cap':    r['end_cap'],
                    'max_dd':     r['max_dd'],
                    'in_portfolio': r['filename'] in portfolio_set,
                }
                for r in single_results
            ],
        }, f, indent=2)
    print(f"{GREEN}Optimales Portfolio in '{OPT_RESULTS}' gespeichert.{NC}")

    # ── Chart & Excel anbieten ────────────────────────────────────────────────
    if final_sim:
        # Nur Portfolio-Strategien in Chart & Excel
        port_single = [r for r in single_results if r['filename'] in portfolio_set]
        if auto:
            # Im Auto-Modus immer Chart + Excel + Telegram (kein Prompt)
            print()
            _generate_portfolio_chart(final_sim, portfolio, capital, start_date, end_date,
                                      port_single)
            _generate_trades_excel(final_sim, portfolio, capital, port_single)
        else:
            print()
            ans = input("  Interaktive Charts & Excel fuer dieses Portfolio erstellen"
                        " & via Telegram senden? (j/n) [Standard: n]: ").strip().lower()
            if ans in ('j', 'y', 'ja'):
                print()
                _generate_portfolio_chart(final_sim, portfolio, capital, start_date, end_date,
                                          port_single)
                _generate_trades_excel(final_sim, portfolio, capital, port_single)


# ---------------------------------------------------------------------------
# Modus 4: Live Signal-Check
# ---------------------------------------------------------------------------

def run_signal_check(symbol: str, timeframe: str):
    from fibot.strategy.fibonacci_logic import generate_signal, signal_summary
    from fibot.analysis.backtester import load_ohlcv, auto_days_for_timeframe

    today      = date.today().isoformat()
    n_days     = min(auto_days_for_timeframe(timeframe), 200)
    start_date = (pd.Timestamp(today, tz='UTC') - pd.Timedelta(days=n_days)).strftime('%Y-%m-%d')

    print(f"\n{CYAN}Lade aktuelle Daten: {symbol} ({timeframe})...{NC}")
    df = load_ohlcv(symbol, timeframe, start_date, today)
    if df.empty:
        print(f"{RED}Keine Daten geladen.{NC}")
        return

    safe     = f"{symbol.replace('/', '').replace(':', '')}_{timeframe}"
    cfg_path = os.path.join(PROJECT_ROOT, 'src', 'fibot', 'strategy', 'configs',
                            f"config_{safe}_fib.json")
    config   = json.load(open(cfg_path)) if os.path.exists(cfg_path) \
               else _default_config(symbol, timeframe)

    print(f"  {len(df)} Kerzen | Letzte Kerze: {df.index[-1]}")
    print(f"\n{CYAN}Berechne Fibonacci-Signal...{NC}\n")

    signal = generate_signal(df, config)
    summary = signal_summary(signal, symbol, timeframe)

    if signal.direction == "none":
        print(f"{YELLOW}{summary}{NC}")
    elif signal.direction == "long":
        print(f"{GREEN}{summary}{NC}")
    else:
        print(f"{RED}{summary}{NC}")


# ---------------------------------------------------------------------------
# Portfolio-Chart & Excel (Mode 3 Post-Optmierung)
# ---------------------------------------------------------------------------

def _get_telegram_cfg() -> tuple[str, str]:
    """Liest Telegram bot_token und chat_id aus secret.json."""
    secret_path = os.path.join(PROJECT_ROOT, 'secret.json')
    try:
        with open(secret_path) as f:
            secrets = json.load(f)
        tg = secrets.get('telegram', {})
        return tg.get('bot_token', ''), tg.get('chat_id', '')
    except Exception:
        return '', ''


def _generate_portfolio_chart(final_sim: dict, portfolio: list,
                               capital: float, start_date: str, end_date: str,
                               all_single_results: list = None):
    """Erstellt einen interaktiven Portfolio-Equity-Chart und sendet ihn optional via Telegram."""
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        print(f"  {RED}plotly nicht installiert — Chart übersprungen. (pip install plotly){NC}")
        return

    eq_df         = final_sim.get('equity_curve')
    trade_history = final_sim.get('trade_history', [])
    if eq_df is None or eq_df.empty:
        print(f"  {YELLOW}Keine Equity-Daten — Chart übersprungen.{NC}")
        return

    portfolio_fnames = {r['filename'] for r in portfolio}

    eq_times = eq_df['timestamp'].astype(str).tolist()
    eq_vals  = eq_df['equity'].tolist()

    # Trade-Marker (nur Portfolio-Trades)
    win_x, win_y   = [], []
    loss_x, loss_y = [], []
    for t in trade_history:
        ts  = str(t['ts'])
        row = eq_df[eq_df['timestamp'] <= t['ts']]
        eq_at = float(row['equity'].iloc[-1]) if not row.empty else capital
        if t['pnl'] > 0:
            win_x.append(ts);  win_y.append(eq_at)
        else:
            loss_x.append(ts); loss_y.append(eq_at)

    n_strats  = len(portfolio)
    pairs_str = ', '.join(
        f"{r['symbol'].split('/')[0]}/{r['timeframe']}" for r in portfolio
    )
    pnl_pct = final_sim['total_pnl_pct']
    sign    = '+' if pnl_pct >= 0 else ''
    title = (
        f"FiBot Portfolio — {n_strats} Strategie(n) ({pairs_str}) | "
        f"Zeitraum: {start_date} → {end_date} | "
        f"Trades: {final_sim['trade_count']} | WR: {final_sim['win_rate']:.1f}% | "
        f"PnL: {sign}{pnl_pct:.1f}% | "
        f"Endkapital: {final_sim['end_capital']:.2f} USDT | "
        f"MaxDD: {final_sim['max_drawdown_pct']:.1f}%"
    )

    fig = make_subplots(specs=[[{"secondary_y": False}]])

    # Startkapital-Referenzlinie
    fig.add_hline(
        y=capital,
        line=dict(color='rgba(100,100,100,0.35)', width=1, dash='dash'),
        annotation_text=f'Start {capital:.0f} USDT',
        annotation_position='top left',
    )

    # Einzelne Equity-Linien für ALLE Strategien (gedimmt im Hintergrund)
    STRAT_COLORS = [
        '#f59e0b', '#10b981', '#8b5cf6', '#f97316',
        '#ec4899', '#14b8a6', '#a3e635', '#fb923c',
    ]
    if all_single_results:
        for idx, sr in enumerate(all_single_results):
            trades = [t for t in sr.get('trade_objs', []) if t.result != 'open']
            if not trades:
                continue
            eq  = capital
            xs, ys = [str(trades[0].timestamp)], [capital]
            for t in trades:
                eq += t.pnl_usdt
                xs.append(str(t.timestamp))
                ys.append(round(eq, 4))
            in_port   = sr['filename'] in portfolio_fnames
            color     = STRAT_COLORS[idx % len(STRAT_COLORS)]
            dash_style = 'solid' if in_port else 'dot'
            opacity   = 0.7 if in_port else 0.35
            label = (f"{sr['symbol'].split('/')[0]}/{sr['timeframe']} "
                     f"({'✔' if in_port else '—'})")
            fig.add_trace(go.Scatter(
                x=xs, y=ys,
                mode='lines', name=label,
                line=dict(color=color, width=1.5, dash=dash_style),
                opacity=opacity,
                hovertemplate=f"{label}: %{{y:.2f}} USDT<extra></extra>",
            ))

    # Portfolio-Equity-Gesamtlinie (vordergrund, blau, fett)
    fig.add_trace(go.Scatter(
        x=eq_times, y=eq_vals,
        mode='lines', name='Portfolio (gesamt)',
        line=dict(color='#2563eb', width=3),
        hovertemplate='Portfolio: %{y:.2f} USDT<extra></extra>',
    ))

    # WIN-Marker ● cyan
    if win_x:
        fig.add_trace(go.Scatter(
            x=win_x, y=win_y, mode='markers',
            marker=dict(color='#22d3ee', symbol='circle', size=8,
                        line=dict(width=1, color='#0e7490')),
            name='TP ✓',
        ))

    # LOSS-Marker ✗ rot
    if loss_x:
        fig.add_trace(go.Scatter(
            x=loss_x, y=loss_y, mode='markers',
            marker=dict(color='#ef4444', symbol='x', size=8,
                        line=dict(width=2, color='#7f1d1d')),
            name='SL ✗',
        ))

    fig.update_layout(
        title=dict(text=title, font=dict(size=12), x=0.5, xanchor='center'),
        height=600,
        hovermode='x unified',
        template='plotly_dark',
        dragmode='zoom',
        xaxis=dict(rangeslider=dict(visible=True), fixedrange=False),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
        margin=dict(l=60, r=60, t=80, b=40),
        yaxis=dict(title='Equity (USDT)', fixedrange=False),
    )

    os.makedirs(os.path.join(PROJECT_ROOT, 'artifacts', 'charts'), exist_ok=True)
    out_file = os.path.join(PROJECT_ROOT, 'artifacts', 'charts', 'fibot_portfolio_equity.html')
    fig.write_html(out_file)
    print(f"  {GREEN}✓ Portfolio-Chart erstellt: {out_file}{NC}")

    bot_token, chat_id = _get_telegram_cfg()
    if bot_token and chat_id:
        from fibot.utils.telegram import send_document
        caption = (
            f"FiBot Portfolio-Equity\n"
            f"{start_date} → {end_date} | {n_strats} Strategie(n) | "
            f"PnL: {sign}{pnl_pct:.1f}% | Equity: {final_sim['end_capital']:.2f} USDT | "
            f"MaxDD: {final_sim['max_drawdown_pct']:.1f}%"
        )
        send_document(bot_token, chat_id, out_file, caption=caption)
        print(f"  {GREEN}✓ Via Telegram gesendet.{NC}")
    else:
        print(f"  {YELLOW}Telegram nicht konfiguriert — Chart nur lokal gespeichert.{NC}")


def _generate_trades_excel(final_sim: dict, portfolio: list, capital: float,
                           all_single_results: list = None):
    """Erstellt eine Excel-Tabelle mit Trades ALLER evaluierten Strategien."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"  {YELLOW}openpyxl nicht installiert — Excel übersprungen. (pip install openpyxl){NC}")
        return

    portfolio_fnames = {r['filename'] for r in portfolio}

    # Alle Trades aus single_results zusammenführen (alle Strategien)
    all_trades = []  # (timestamp, fname, symbol, timeframe, trade_obj, in_portfolio)
    if all_single_results:
        for sr in all_single_results:
            in_port = sr['filename'] in portfolio_fnames
            for t in sr.get('trade_objs', []):
                if t.result == 'open':
                    continue
                all_trades.append((t.timestamp, sr['filename'], sr['symbol'],
                                   sr['timeframe'], t, in_port))
        all_trades.sort(key=lambda x: x[0])
    else:
        # Fallback: nur Portfolio-Sim-Trades
        for t in final_sim.get('trade_history', []):
            all_trades.append((t['ts'], t['fname'], '', '', t, True))

    if not all_trades:
        print(f"  {YELLOW}Keine Trades — Excel übersprungen.{NC}")
        return

    equity = capital
    rows = []
    for i, (ts, fname, symbol, timeframe, t, in_port) in enumerate(all_trades):
        if isinstance(t, dict):
            pnl  = float(t['pnl'])
            dir_ = t['direction'].upper()
            entr = round(float(t['entry']), 6)
            ex   = round(float(t['exit']),  6)
            ergebnis = 'TP erreicht' if pnl > 0 else 'SL erreicht'
        else:
            pnl  = float(t.pnl_usdt)
            dir_ = t.direction.upper()
            entr = round(float(t.entry),       6)
            ex   = round(float(t.exit_price),  6)
            ergebnis = 'TP erreicht' if t.result == 'win' else 'SL erreicht'
        equity += pnl
        fname_short = fname.replace('config_', '').replace('_fib.json', '')
        rows.append({
            'Nr':            i + 1,
            'Datum':         str(ts)[:16].replace('T', ' '),
            'Strategie':     fname_short,
            'Portfolio':     '✔' if in_port else '—',
            'Richtung':      dir_,
            'Entry':         entr,
            'Exit':          ex,
            'Ergebnis':      ergebnis,
            'PnL (USDT)':    round(pnl,    4),
            'Kapital':       round(equity, 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'

    header_fill  = PatternFill('solid', fgColor='1E3A5F')
    win_fill     = PatternFill('solid', fgColor='D6F4DC')
    loss_fill    = PatternFill('solid', fgColor='FAD7D7')
    alt_fill     = PatternFill('solid', fgColor='F2F2F2')
    thin_border  = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )
    col_widths = {
        'Nr': 5, 'Datum': 18, 'Strategie': 28, 'Portfolio': 10, 'Richtung': 10,
        'Entry': 14, 'Exit': 14, 'Ergebnis': 14, 'PnL (USDT)': 14, 'Kapital': 16,
    }

    headers = list(rows[0].keys())
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 14)
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(rows, 2):
        fill = win_fill if row['Ergebnis'] == 'TP erreicht' else \
               loss_fill if r_idx % 2 == 0 else alt_fill
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=col, value=row[key])
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Entry', 'Exit', 'PnL (USDT)', 'Kapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[r_idx].height = 18

    # Zusammenfassung
    total_trades = len(rows)
    wins = sum(1 for r in rows if r['Ergebnis'] == 'TP erreicht')
    sr = total_trades + 3
    ws.cell(row=sr, column=1, value='Zusammenfassung').font = Font(bold=True, size=11)
    pnl_total = rows[-1]['Kapital'] - capital if rows else 0.0
    pnl_pct   = pnl_total / capital * 100 if capital else 0.0
    for label, value in [
        ('Trades gesamt',  total_trades),
        ('Win-Rate',       f"{wins / total_trades * 100:.1f}%" if total_trades else '—'),
        ('PnL',            f"{pnl_pct:+.1f}%"),
        ('Endkapital',     f"{rows[-1]['Kapital']:.2f} USDT" if rows else '—'),
    ]:
        ws.cell(row=sr, column=1, value=label).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=value)
        sr += 1

    os.makedirs(os.path.join(PROJECT_ROOT, 'artifacts', 'charts'), exist_ok=True)
    out_file = os.path.join(PROJECT_ROOT, 'artifacts', 'charts', 'fibot_trades.xlsx')
    wb.save(out_file)
    print(f"  {GREEN}✓ Excel-Tabelle erstellt: {out_file}{NC}")

    bot_token, chat_id = _get_telegram_cfg()
    if bot_token and chat_id:
        from fibot.utils.telegram import send_document
        send_document(bot_token, chat_id, out_file,
                      caption=f"FiBot Trades — {total_trades} Trades, "
                              f"WR: {wins / total_trades * 100:.1f}%" if total_trades else "FiBot Trades")
        print(f"  {GREEN}✓ Via Telegram gesendet.{NC}")


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

def _print_result(result, compact: bool = False):
    from fibot.analysis.backtester import BacktestResult
    pnl_color = GREEN if result.pnl_pct >= 0 else RED

    if compact:
        print(f"  Kapital: {result.start_capital:.0f} → {result.end_capital:.2f} USDT "
              f"({pnl_color}{result.pnl_pct:+.2f}%{NC}) | "
              f"Trades: {result.total_trades} | WR: {result.win_rate:.1f}% | "
              f"MaxDD: {result.max_drawdown_pct:.2f}% | Avg R:R 1:{result.avg_rr:.2f}")
        return

    print(f"{'═'*55}")
    print(f"{BOLD}Backtest: {result.symbol} ({result.timeframe}){NC}")
    print(f"{'─'*55}")
    print(f"  Kapital     : {result.start_capital:.2f} → "
          f"{pnl_color}{result.end_capital:.2f} USDT{NC} "
          f"({pnl_color}{result.pnl_pct:+.2f}%{NC})")
    print(f"  Trades      : {result.total_trades}  "
          f"({GREEN}W:{result.wins}{NC} / {RED}L:{result.losses}{NC})")
    print(f"  Win-Rate    : {result.win_rate:.1f}%")
    print(f"  Max Drawdown: {result.max_drawdown_pct:.2f}%")
    print(f"  Avg R:R     : 1:{result.avg_rr:.2f}")

    if result.trades:
        closed = [t for t in result.trades if t.result != 'open']
        if closed:
            longs  = len([t for t in closed if t.direction == 'long'])
            shorts = len([t for t in closed if t.direction == 'short'])
            print(f"  Long/Short  : {longs} / {shorts}")
            avg_hold = sum(t.hold_bars for t in closed) / len(closed)
            print(f"  Ø Haltedauer: {avg_hold:.1f} Kerzen")
    print(f"{'═'*55}")


def _print_json_result(d: dict):
    pnl_color = GREEN if d.get('pnl_pct', 0) >= 0 else RED
    print(f"\n{'═'*55}")
    print(f"{BOLD}{d['symbol']} ({d['timeframe']}){NC}")
    print(f"{'─'*55}")
    print(f"  Kapital    : {d['start_capital']:.2f} → "
          f"{pnl_color}{d['end_capital']:.2f} USDT{NC} "
          f"({pnl_color}{d['pnl_pct']:+.2f}%{NC})")
    print(f"  Trades     : {d['total_trades']}  (W:{d['wins']} / L:{d['losses']})")
    print(f"  Win-Rate   : {d['win_rate']:.1f}%")
    print(f"  Max DD     : {d['max_drawdown']:.2f}%")
    print(f"  Avg R:R    : 1:{d['avg_rr']:.2f}")

    trades = d.get('trades', [])
    if trades:
        print(f"\n  Letzte 5 Trades:")
        print(f"  {'Datum':<22} {'Dir':<7} {'Entry':>10} {'Exit':>10} {'PnL':>9} {'Erg'}")
        print(f"  {'─'*65}")
        for t in trades[-5:]:
            color  = GREEN if t['result'] == 'win' else (RED if t['result'] == 'loss' else NC)
            result_str = '✓' if t['result'] == 'win' else ('✗' if t['result'] == 'loss' else '…')
            print(f"  {t['ts'][:19]:<22} {t['direction']:<7} "
                  f"{t['entry']:>10.4f} {t['exit']:>10.4f} "
                  f"{color}{t['pnl_usdt']:>+8.2f}$  {result_str}{NC}")
    print(f"{'═'*55}\n")


def _default_config(symbol: str, timeframe: str) -> dict:
    return {
        "market":   {"symbol": symbol, "timeframe": timeframe},
        "strategy": {
            "swing_lookback": 100, "pivot_left": 5, "pivot_right": 5,
            "structure_lookback": 60, "fib_entry_min": 0.382, "fib_entry_max": 0.618,
            "fib_sl_level": 0.786, "fib_tp1_level": 1.000, "fib_tp2_level": 1.272,
            "fib_tolerance_atr_mult": 0.5, "structure_tolerance_atr_mult": 0.3,
            "rsi_period": 14, "rsi_oversold": 45, "rsi_overbought": 55,
            "volume_ratio_min": 1.0, "min_rr": 1.5, "atr_period": 14,
            "atr_sl_multiplier": 1.5, "min_signal_score": 4.0, "candle_limit": 500,
        },
        "risk": {"leverage": 10, "risk_per_entry_pct": 1.0, "margin_mode": "isolated"},
    }


# ---------------------------------------------------------------------------
# CLI (wird von show_results.sh aufgerufen)
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="FiBot Show Results")
    parser.add_argument('--mode',          type=int, required=True, help="1-4")
    parser.add_argument('--from',          dest='date_from', default=None)
    parser.add_argument('--to',            dest='date_to',   default=None)
    parser.add_argument('--capital',       type=float, default=1000.0)
    parser.add_argument('--configs',       default=None,
                        help="Leerzeichen-getrennte Config-Dateinamen für Modus 2")
    parser.add_argument('--target-max-dd', type=float, default=30.0,
                        help="Max Drawdown %% für Portfolio-Finder (Modus 3)")
    parser.add_argument('--min-wr',        type=float, default=0.0,
                        help="Min Win-Rate %% für Portfolio-Finder (Modus 3)")
    parser.add_argument('--auto',          action='store_true', default=False,
                        help="Nicht-interaktiver Modus (Auto-Optimizer): überspringt alle Prompts")
    parser.add_argument('--symbols',       default=None,
                        help="Nur diese Coins optimieren, z.B. 'BTC ETH ADA' (Modus 3)")
    args = parser.parse_args()

    today = date.today().isoformat()
    start = args.date_from if args.date_from else '2024-01-01'
    end   = args.date_to   if args.date_to   else today

    if args.mode == 1:
        run_all_configs_isolated(start, end, args.capital)

    elif args.mode == 2:
        if not args.configs:
            print(f"{RED}--configs erforderlich für Modus 2.{NC}")
            sys.exit(1)
        filenames = args.configs.split()
        run_manual_portfolio(filenames, start, end, args.capital)

    elif args.mode == 3:
        symbols = args.symbols.split() if args.symbols else None
        configs = args.configs.split() if args.configs else None
        run_portfolio_finder(args.capital, args.target_max_dd, args.min_wr, start, end,
                             auto=args.auto, symbols=symbols, configs=configs)

    elif args.mode == 4:
        from fibot.analysis.interactive_chart import run_interactive_chart
        secret_path = os.path.join(PROJECT_ROOT, 'secret.json')
        secrets = {}
        if os.path.exists(secret_path):
            with open(secret_path) as f:
                secrets = json.load(f)
        run_interactive_chart(secrets)

    else:
        print(f"{RED}Ungültiger Modus.{NC}")
        sys.exit(1)

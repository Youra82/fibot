#!/usr/bin/env python3
"""
run_portfolio_optimizer.py  (fibot)

Lädt alle Configs, führt Portfolio-Simulation (gemeinsamer Kapital-Pool,
kombinierte Equity-Kurve, echter MaxDD) durch und wählt das beste Portfolio
per Calmar-Greedy. Schreibt active_strategies in settings.json.

Aufruf:
  python3 run_portfolio_optimizer.py              # interaktiv
  python3 run_portfolio_optimizer.py --auto-write # automatisch (Scheduler)
"""
import contextlib
import io
import os
import sys
import json
import argparse
from datetime import date, timedelta
from tqdm import tqdm

PROJECT_ROOT  = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))

CONFIGS_DIR   = os.path.join(PROJECT_ROOT, 'src', 'fibot', 'strategy', 'configs')
SETTINGS_PATH = os.path.join(PROJECT_ROOT, 'settings.json')

B  = '\033[1;37m'
G  = '\033[0;32m'
Y  = '\033[1;33m'
R  = '\033[0;31m'
NC = '\033[0m'

DEFAULT_LOOKBACK_DAYS = 1095  # ~3 Jahre als Standard


def _calmar(pnl_pct: float, max_dd_pct: float) -> float:
    return pnl_pct / max_dd_pct if max_dd_pct > 0 else pnl_pct


def _scan_configs() -> list:
    if not os.path.isdir(CONFIGS_DIR):
        return []
    return sorted([
        os.path.join(CONFIGS_DIR, f)
        for f in os.listdir(CONFIGS_DIR)
        if f.endswith('.json')
    ])


def _build_strategies_data(config_files: list, start_date: str, end_date: str) -> dict:
    from fibot.analysis.backtester import load_ohlcv
    strategies_data = {}
    for path in tqdm(config_files, desc='Lade Configs & Daten'):
        fname = os.path.basename(path)
        try:
            with open(path) as f:
                config = json.load(f)
            market    = config.get('market', {})
            symbol    = market.get('symbol', '')
            timeframe = market.get('timeframe', '')
            if not symbol or not timeframe:
                continue
            df = load_ohlcv(symbol, timeframe, start_date, end_date)
            if df is None or df.empty or len(df) < 50:
                print(f"  {Y}Uebersprungen (keine Daten): {fname}{NC}")
                continue
            strategies_data[fname] = {
                'symbol':    symbol,
                'timeframe': timeframe,
                'df':        df,
                'config':    config,
            }
        except Exception as e:
            print(f"  {Y}Fehler bei {fname}: {e}{NC}")
    return strategies_data


def _simulate_silent(start_capital: float, subset: dict,
                     start_date: str, end_date: str) -> dict | None:
    from fibot.analysis.portfolio_simulator import run_portfolio_simulation
    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        return run_portfolio_simulation(start_capital, subset, start_date, end_date)


def _run_portfolio_optimizer(start_capital: float, strategies_data: dict,
                              start_date: str, end_date: str,
                              target_max_dd: float) -> dict:
    print(f"\n--- Portfolio-Optimierung: Max DD <= {target_max_dd:.2f}% ---")
    filenames = list(strategies_data.keys())
    total     = len(filenames)

    # 1. Pre-Filter: Einzeln simulieren
    print(f"1/2: Pre-Filter — {total} Configs werden einzeln getestet...")
    valid_candidates = []
    for i, fname in enumerate(filenames, 1):
        print(f"\r  [{i:>3}/{total}] {fname:<50}", end='', flush=True)
        subset = {fname: strategies_data[fname]}
        res = _simulate_silent(start_capital, subset, start_date, end_date)
        if not res or res.get('liquidation_date'):
            continue
        max_dd = res.get('max_drawdown_pct', 100.0)
        pnl    = res.get('total_pnl_pct', 0.0)
        if max_dd <= target_max_dd and pnl > 0:
            valid_candidates.append({
                'fname':       fname,
                'end_capital': res['end_capital'],
                'max_dd':      max_dd,
                'pnl_pct':     pnl,
                'result':      res,
            })
    print()

    if not valid_candidates:
        print(f"Keine Einzelstrategie erfuellte Max DD <= {target_max_dd:.2f}%.")
        return {'optimal_portfolio': [], 'final_result': None}

    valid_candidates.sort(key=lambda c: _calmar(c['pnl_pct'], c['max_dd']), reverse=True)
    print(f"-> {len(valid_candidates)}/{total} Kandidaten bestehen Pre-Filter:")
    for c in valid_candidates:
        print(f"   {c['fname']:<50} | Kapital: {c['end_capital']:>8.2f} USDT | DD: {c['max_dd']:>5.1f}%")

    # 2. Greedy-Selektion (Calmar-optimierend, Coin-Kollision vermeiden)
    print(f"\n2/2: Greedy-Selektion...")
    candidate_files = [c['fname'] for c in valid_candidates]
    best_files      = [candidate_files[0]]
    best_result     = valid_candidates[0]['result']
    best_calmar     = _calmar(valid_candidates[0]['pnl_pct'], valid_candidates[0]['max_dd'])
    selected_coins  = {strategies_data[candidate_files[0]]['symbol'].split('/')[0]}
    remaining       = candidate_files[1:]

    step = 0
    while remaining:
        step += 1
        best_addition        = None
        best_addition_calmar = best_calmar
        best_addition_result = best_result

        candidates_this_round = [
            f for f in remaining
            if strategies_data[f]['symbol'].split('/')[0] not in selected_coins
        ]
        for idx, candidate in enumerate(candidates_this_round):
            team   = best_files + [candidate]
            subset = {f: strategies_data[f] for f in team}
            print(f"\r  Schritt {step} | Kandidat {idx+1}/{len(candidates_this_round)} | "
                  f"Beste Calmar: {best_calmar:.2f}", end='', flush=True)
            res = _simulate_silent(start_capital, subset, start_date, end_date)
            if not res or res.get('liquidation_date'):
                continue
            if res.get('max_drawdown_pct', 100.0) > target_max_dd:
                continue
            calmar = _calmar(res.get('total_pnl_pct', 0.0), res.get('max_drawdown_pct', 100.0))
            if calmar > best_addition_calmar:
                best_addition_calmar = calmar
                best_addition        = candidate
                best_addition_result = res

        if best_addition:
            best_files.append(best_addition)
            selected_coins.add(strategies_data[best_addition]['symbol'].split('/')[0])
            best_calmar = best_addition_calmar
            best_result = best_addition_result
            remaining.remove(best_addition)
            sd = strategies_data[best_addition]
            print(f"\n  + {sd['symbol']} {sd['timeframe']} "
                  f"(Calmar: {best_calmar:.2f} | "
                  f"PnL: {best_result['total_pnl_pct']:+.1f}% | "
                  f"DD: {best_result['max_drawdown_pct']:.1f}%)")
        else:
            print(f"\n  Keine weitere Verbesserung. Optimierung beendet.")
            break

    return {'optimal_portfolio': best_files, 'final_result': best_result}


def _simulate_current_portfolio(settings: dict, strategies_data: dict,
                                 start_capital: float,
                                 start_date: str, end_date: str) -> dict | None:
    """Simuliert das aktuell aktive Portfolio auf dem gleichen Zeitraum."""
    current = [
        s for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
        if s.get('active')
    ]
    if not current:
        return None
    subset = {}
    for s in current:
        sym, tf = s.get('symbol', ''), s.get('timeframe', '')
        for fname, sd in strategies_data.items():
            if sd['symbol'] == sym and sd['timeframe'] == tf:
                subset[fname] = sd
                break
    if not subset:
        return None
    return _simulate_silent(start_capital, subset, start_date, end_date)


def _write_to_settings(portfolio_files: list, strategies_data: dict) -> None:
    with open(SETTINGS_PATH) as f:
        settings = json.load(f)
    existing     = settings.get('live_trading_settings', {}).get('active_strategies', [])
    existing_map = {(s.get('symbol'), s.get('timeframe')): s for s in existing}
    new_strategies = []
    for fname in portfolio_files:
        sd        = strategies_data.get(fname, {})
        symbol    = sd.get('symbol', '')
        timeframe = sd.get('timeframe', '')
        if not symbol or not timeframe:
            continue
        config = sd.get('config', {})
        risk   = config.get('risk', {})
        base   = existing_map.get((symbol, timeframe), {})
        entry  = {
            **base,
            'symbol':             symbol,
            'timeframe':          timeframe,
            'leverage':           risk.get('leverage',           base.get('leverage',           2)),
            'margin_mode':        risk.get('margin_mode',        base.get('margin_mode',        'isolated')),
            'risk_per_entry_pct': risk.get('risk_per_entry_pct', base.get('risk_per_entry_pct', 0.5)),
            'active':             True,
        }
        new_strategies.append(entry)
    lt = settings.setdefault('live_trading_settings', {})
    lt['active_strategies']          = new_strategies
    lt['use_auto_optimizer_results'] = True
    with open(SETTINGS_PATH, 'w') as f:
        json.dump(settings, f, indent=2)


BOT_NAME = 'fibot'


def _get_telegram_creds():
    try:
        with open(os.path.join(PROJECT_ROOT, 'secret.json')) as f:
            s = json.load(f)
        tg = s.get('telegram', {})
        t, c = tg.get('bot_token', ''), tg.get('chat_id', '')
        return (t, c) if t and c else (None, None)
    except Exception:
        return None, None


def _send_telegram(msg):
    token, chat = _get_telegram_creds()
    if not token:
        return
    try:
        import requests
        requests.post(f'https://api.telegram.org/bot{token}/sendMessage',
                      data={'chat_id': chat, 'text': msg}, timeout=10)
    except Exception:
        pass


def _send_telegram_doc(fpath, caption=''):
    token, chat = _get_telegram_creds()
    if not token:
        return
    try:
        import requests
        with open(fpath, 'rb') as fh:
            requests.post(f'https://api.telegram.org/bot{token}/sendDocument',
                          data={'chat_id': chat, 'caption': caption},
                          files={'document': fh}, timeout=30)
    except Exception:
        pass


def generate_trades_excel(final, strategies_data, capital, start_date, end_date):
    """Erstellt Excel-Tabelle mit allen Portfolio-Trades."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f'  {Y}openpyxl nicht installiert — Excel uebersprungen.{NC}')
        return None

    trades = final.get('trade_history', [])
    if not trades:
        return None

    equity = capital
    rows = []
    for i, t in enumerate(trades, 1):
        fname  = t.get('fname', '')
        sd     = strategies_data.get(fname, {})
        symbol = sd.get('symbol', fname.split('.')[0] if fname else '?')
        tf     = sd.get('timeframe', '?')
        pnl    = t.get('pnl', 0.0)
        equity += pnl
        rows.append({
            'Nr':            i,
            'Datum':         str(t.get('ts', ''))[:16].replace('T', ' '),
            'Symbol':        symbol,
            'Timeframe':     tf,
            'Richtung':      str(t.get('direction', '?')).upper(),
            'Ergebnis':      'TP erreicht' if pnl >= 0 else 'SL erreicht',
            'PnL (USDT)':    round(pnl, 4),
            'Gesamtkapital': round(equity, 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'
    hdr  = PatternFill('solid', fgColor='1E3A5F')
    win  = PatternFill('solid', fgColor='D6F4DC')
    loss = PatternFill('solid', fgColor='FAD7D7')
    alt  = PatternFill('solid', fgColor='F2F2F2')
    brd  = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                  top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
    cw   = {'Nr': 6, 'Datum': 18, 'Symbol': 22, 'Timeframe': 12, 'Richtung': 10,
             'Ergebnis': 14, 'PnL (USDT)': 14, 'Gesamtkapital': 16}
    hdrs = list(rows[0].keys())
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = brd
        ws.column_dimensions[get_column_letter(c)].width = cw.get(h, 14)
    ws.row_dimensions[1].height = 22
    for ri, row in enumerate(rows, 2):
        f = win if row['Ergebnis'] == 'TP erreicht' else (loss if ri % 2 == 0 else alt)
        for c, key in enumerate(hdrs, 1):
            cell = ws.cell(row=ri, column=c, value=row[key])
            cell.fill = f
            cell.border = brd
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('PnL (USDT)', 'Gesamtkapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[ri].height = 18
    pnl = final.get('total_pnl_pct', 0)
    dd  = final.get('max_drawdown_pct', 0)
    wr  = final.get('win_rate', 0)
    eq  = final.get('end_capital', equity)
    n   = final.get('trade_count', len(trades))
    sr  = len(rows) + 3
    for label, val in [('Zeitraum', f'{start_date} -> {end_date}'), ('Trades', n),
                        ('Win-Rate', f'{wr:.1f}%'), ('PnL', f'{pnl:+.1f}%'),
                        ('Endkapital', f'{eq:.2f} USDT'), ('Max Drawdown', f'{dd:.1f}%')]:
        ws.cell(row=sr, column=1, value=label).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=val)
        sr += 1
    outfile = f'/tmp/{BOT_NAME}_trades.xlsx'
    wb.save(outfile)
    print(f'  {G}✓ Excel erstellt: {outfile}{NC}')
    return outfile


def generate_equity_html(final, capital, start_date, end_date, labels):
    """Erstellt interaktiven Portfolio-Equity-Chart."""
    try:
        import plotly.graph_objects as go
    except ImportError:
        print(f'  {Y}plotly nicht installiert — Chart uebersprungen.{NC}')
        return None

    eq_df = final.get('equity_curve')
    if eq_df is None or (hasattr(eq_df, 'empty') and eq_df.empty):
        return None

    times = [str(t) for t in eq_df['timestamp']]
    vals  = [float(v) for v in eq_df['equity']]
    pnl   = final.get('total_pnl_pct', 0)
    dd    = final.get('max_drawdown_pct', 0)
    wr    = final.get('win_rate', 0)
    n     = final.get('trade_count', 0)
    eq    = final.get('end_capital', vals[-1] if vals else capital)
    sign  = '+' if pnl >= 0 else ''
    title = (f"{BOT_NAME} Portfolio — {', '.join(labels)} | "
             f"PnL: {sign}{pnl:.1f}% | Equity: {eq:.2f} USDT | "
             f"MaxDD: {dd:.1f}% | WR: {wr:.1f}% | {n} Trades")

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=times, y=vals, mode='lines', name='Portfolio Equity',
                             line=dict(color='#2563eb', width=2)))
    fig.add_hline(y=capital, line=dict(color='rgba(100,100,100,0.4)', width=1, dash='dash'),
                  annotation_text=f'Start {capital:.0f} USDT', annotation_position='top left')
    fig.update_layout(title=dict(text=title, font=dict(size=12), x=0.5),
                      height=600, template='plotly_white', hovermode='x unified',
                      xaxis=dict(rangeslider=dict(visible=True), fixedrange=False),
                      yaxis=dict(title='Equity (USDT)', fixedrange=False))
    outfile = f'/tmp/{BOT_NAME}_portfolio_equity.html'
    fig.write_html(outfile)
    print(f'  {G}✓ Chart erstellt: {outfile}{NC}')
    return outfile


def _do_replot(settings: dict, capital: float, start_date: str, end_date: str) -> int:
    print(f"\n{'─'*72}")
    print(f"{B}  fibot — Replot (aktives Portfolio){NC}")
    print(f"  Kapital: {capital:.0f} USDT | Zeitraum: {start_date} → {end_date}")
    print(f"{'─'*72}\n")

    active = [s for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
              if s.get('active')]
    if not active:
        print(f"{R}  Keine aktiven Strategien in settings.json.{NC}")
        return 1

    active_pairs = {(s['symbol'], s['timeframe']) for s in active}
    matching = []
    for path in _scan_configs():
        try:
            with open(path) as f:
                cfg = json.load(f)
            m = cfg.get('market', {})
            if (m.get('symbol'), m.get('timeframe')) in active_pairs:
                matching.append(path)
        except Exception:
            pass

    if not matching:
        print(f"{R}  Keine Config-Dateien fuer aktive Strategien gefunden.{NC}")
        return 1

    print(f"  {len(matching)} Config(s) gefunden.\n")
    strategies_data = _build_strategies_data(matching, start_date, end_date)
    if not strategies_data:
        print(f"{R}  Keine Daten geladen.{NC}")
        return 1

    final = _simulate_silent(capital, strategies_data, start_date, end_date)
    if not final:
        print(f"{R}  Portfolio-Simulation fehlgeschlagen.{NC}")
        return 1

    selected_files = list(strategies_data.keys())
    labels = [f"{sd.get('symbol', '?')}/{sd.get('timeframe', '?')}"
              for sd in strategies_data.values()]
    pnl = final.get('total_pnl_pct', 0)
    dd  = final.get('max_drawdown_pct', 0)
    n   = final.get('trade_count', 0)
    wr  = final.get('win_rate', 0)
    eq  = final.get('end_capital', 0)

    print(f"\n{'='*72}")
    print(f"{B}  Replot — {len(selected_files)} Strategie(n){NC}\n")
    for fname, sd in strategies_data.items():
        print(f"  {G}✓{NC} {sd.get('symbol', fname):<26} / {sd.get('timeframe', ''):<6}")
    print(f"\n  Endkapital: {eq:.2f} USDT  | PnL: {pnl:+.1f}%  | MaxDD: {dd:.2f}%")
    print(f"{'='*72}\n")

    summary = (f"{BOT_NAME} Replot\n"
               f"{len(selected_files)} Strategien | {n} Trades | WR: {wr:.1f}%\n"
               f"PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}% | Equity: {eq:.2f} USDT\n"
               f"Zeitraum: {start_date} -> {end_date}")
    _send_telegram(summary)
    xlsx = generate_trades_excel(final, strategies_data, capital, start_date, end_date)
    if xlsx:
        _send_telegram_doc(xlsx, caption=f'{BOT_NAME} Trades | {n} Trades | WR: {wr:.1f}% | Equity: {eq:.2f} USDT')
    html = generate_equity_html(final, capital, start_date, end_date, labels)
    if html:
        _send_telegram_doc(html, caption=f'{BOT_NAME} Portfolio-Equity | PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}%')
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description='fibot Portfolio-Optimizer')
    parser.add_argument('--capital',    type=float, default=None)
    parser.add_argument('--max-dd',     type=float, default=30.0)
    parser.add_argument('--start-date', type=str,   default=None)
    parser.add_argument('--end-date',   type=str,   default=None)
    parser.add_argument('--auto-write', action='store_true')
    parser.add_argument('--replot',     action='store_true',
                        help='Replot fuer aktives Portfolio (keine Re-Optimierung)')
    args = parser.parse_args()

    with open(SETTINGS_PATH) as f:
        settings = json.load(f)
    opt           = settings.get('optimization_settings', {})
    capital       = args.capital or float(opt.get('start_capital', 100))
    max_dd        = args.max_dd
    end_date      = args.end_date   or date.today().strftime('%Y-%m-%d')
    start_date    = args.start_date or (
        date.today() - timedelta(days=DEFAULT_LOOKBACK_DAYS)
    ).strftime('%Y-%m-%d')
    max_positions = int(settings.get('live_trading_settings', {}).get('max_open_positions', 7))

    if args.replot:
        return _do_replot(settings, capital, start_date, end_date)

    print(f"\n{'─'*72}")
    print(f"{B}  fibot — Automatische Portfolio-Optimierung{NC}")
    print(f"  Greedy-Selektion mit echter Portfolio-Simulation (MaxDD ≤ {max_dd:.0f}%)")
    print(f"  Kapital: {capital:.0f} USDT | Positionen: max {max_positions} | "
          f"Zeitraum: {start_date} → {end_date}")
    print(f"{'─'*72}\n")

    config_files = _scan_configs()
    if not config_files:
        print(f"{R}  Keine Configs in {CONFIGS_DIR}{NC}")
        print(f"  → Zuerst run_pipeline.sh ausfuehren!\n")
        return 1

    print(f"  {len(config_files)} Config(s) gefunden.\n")
    strategies_data = _build_strategies_data(config_files, start_date, end_date)
    if not strategies_data:
        print(f"{R}  Keine Daten geladen.{NC}")
        return 1

    result = _run_portfolio_optimizer(capital, strategies_data, start_date, end_date, max_dd)

    if not result or not result.get('optimal_portfolio'):
        print(f"{R}  Kein Portfolio erfuellt die Bedingungen (MaxDD ≤ {max_dd:.0f}%).{NC}\n")
        return 0

    portfolio_files = result['optimal_portfolio'][:max_positions]
    final           = result.get('final_result') or {}

    print(f"\n{'='*72}")
    print(f"{B}  Optimales Portfolio — {len(portfolio_files)} Strategie(n){NC}\n")
    for fname in portfolio_files:
        sd = strategies_data.get(fname, {})
        print(f"  {G}✓{NC} {sd.get('symbol', fname):<26} / {sd.get('timeframe', ''):<6}")
    if final:
        pnl = final.get('total_pnl_pct', 0)
        print(f"\n  Endkapital: {final.get('end_capital', 0):.2f} USDT  "
              f"| PnL: {pnl:+.1f}%  "
              f"| MaxDD: {final.get('max_drawdown_pct', 0):.2f}%")
    print(f"{'='*72}\n")

    current_set = {
        (s.get('symbol'), s.get('timeframe'))
        for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
        if s.get('active')
    }
    new_set = {
        (strategies_data.get(f, {}).get('symbol'), strategies_data.get(f, {}).get('timeframe'))
        for f in portfolio_files
    }

    cur_result  = _simulate_current_portfolio(settings, strategies_data, capital, start_date, end_date)
    cur_cap     = cur_result.get('end_capital', 0) if cur_result else 0
    new_cap     = final.get('end_capital', 0)
    if cur_result:
        print(f"  Aktuelles Portfolio: {cur_cap:.2f} USDT  "
              f"| PnL: {cur_result.get('total_pnl_pct', 0):+.1f}%  "
              f"| MaxDD: {cur_result.get('max_drawdown_pct', 0):.2f}%")
        print(f"  Neues Portfolio:     {new_cap:.2f} USDT  "
              f"| PnL: {final.get('total_pnl_pct', 0):+.1f}%  "
              f"| MaxDD: {final.get('max_drawdown_pct', 0):.2f}%\n")

    if args.auto_write:
        if cur_result and new_cap <= cur_cap:
            print(f"{Y}  Neues Portfolio ({new_cap:.2f} USDT) nicht besser als aktuelles "
                  f"({cur_cap:.2f} USDT) — keine Aenderung.{NC}\n")
        else:
            _write_to_settings(portfolio_files, strategies_data)
            print(f"{G}✓ settings.json aktualisiert — {len(portfolio_files)} Strategie(n).{NC}\n")
    else:
        if current_set == new_set:
            print(f"{Y}  Portfolio unveraendert — keine Aenderung noetig.{NC}\n")
        else:
            try:
                ans = input("  Optimales Portfolio in settings.json eintragen? (j/n): ").strip().lower()
            except (EOFError, KeyboardInterrupt):
                ans = 'n'
            if ans in ('j', 'ja', 'y', 'yes'):
                _write_to_settings(portfolio_files, strategies_data)
                print(f"{G}✓ settings.json aktualisiert.{NC}\n")
            else:
                print(f"{Y}  settings.json NICHT geaendert.{NC}\n")

    # ── Reports & Telegram ──────────────────────────────────────────────────
    if args.auto_write:
        labels = [
            f"{strategies_data.get(f, {}).get('symbol', '?')}/{strategies_data.get(f, {}).get('timeframe', '?')}"
            for f in portfolio_files
        ]
        pnl = final.get('total_pnl_pct', 0)
        dd  = final.get('max_drawdown_pct', 0)
        n   = final.get('trade_count', 0)
        wr  = final.get('win_rate', 0)
        eq  = final.get('end_capital', 0)
        summary = (f"{BOT_NAME} Auto-Optimizer\n"
                   f"{len(portfolio_files)} Strategien | {n} Trades | WR: {wr:.1f}%\n"
                   f"PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}% | Equity: {eq:.2f} USDT\n"
                   f"Zeitraum: {start_date} -> {end_date}")
        _send_telegram(summary)
        xlsx = generate_trades_excel(final, strategies_data, capital, start_date, end_date)
        if xlsx:
            _send_telegram_doc(xlsx, caption=f'{BOT_NAME} Trades | {n} Trades | WR: {wr:.1f}% | Equity: {eq:.2f} USDT')
        html = generate_equity_html(final, capital, start_date, end_date, labels)
        if html:
            _send_telegram_doc(html, caption=f'{BOT_NAME} Portfolio-Equity | PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}%')

    return 0


if __name__ == '__main__':
    sys.exit(main())
